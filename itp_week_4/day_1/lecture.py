# ITP Week 4 Day 1 Lecture

###########################
#
# import packages
#
###########################

import openpyxl
import requests
import json 

###########################
#
# set up openpyl stuff
#
###########################


def retrieve_wb(fileame):
    print("Loading workbook....")
    return openpyxl.load_workbook(filename)

def retrieve_first_ws_from_wb(wb):
    print("Retrieving Worksheet")
    first_ws_name = wb.sheetnames[0]
    return wb[first_ws_name]

#ws = retrieve_first_ws_from_wb(filename)

###########################
#
# set up api call stuff
# get the data and transform it
#
###########################

def return_json_data(url):
    print("Requesting data....")
    r = requests.get(url)
    print("Data successfully retrieved")
    return json.loads(r.text)

###########################
#
# write the data to the worksheet
#
###########################


def write_pokemon_name(data, ws):
    print("Writing pokemon names....")
    ws['A1'] = data['results'][0]['name']
    for i in range(100):
        ws['A' + str(i + 1)] = data['results'][i]['name']
    print("Completed writing.")

###########################
#
# save the worksheet
#
###########################

def save_wb(wb, filename):
    print("Saving workbook...")
    wb.save(filename)
    print("Workbook has been successfully saved")




################### FUNCTION CALL SECTION ##################

filename = "/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_4/day_1/output.xlsx"
url = "https://pokeapi.co/api/v2/pokemon?limit=100"


wb = retrieve_wb(filename)
ws = retrieve_first_ws_from_wb(wb)
data = return_json_data(url)
write_pokemon_name(data, ws)
save_wb(wb, filename)