#------------EXCEL SPREADSHEET-----------

# -Workbook
#     -can contain multiple sheets (worksheets)
# -xlsx file extension
# -Sheets  / Worksheets
# -Columns (letters) & Rows (numbers)
# -Cell
#     -intersection of row & Column

# READ
# - SHEET, CELLS, COLUMNS, ROWS

#------------OPENPYXL-----------
# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).

# -openpyxl is a downloaded module
# -Download & install openpyxl using pip
#   pip install openpyxl

import openpyxl

#Assign your .xlsx file to a variable
filename = "/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_3/day_3/lecture.xlsx"
openpyxl.load_workbook(filename) #returns a Workbook object.
my_workbook = openpyxl.load_workbook('example.xlsx')
#use the 'type' method on the new variable to verify what kind of data type you are working with
type(my_workbook) # Result-->  <class 'openpyxl.workbook.workbook.Workbook'>



#Each Excel file can have multiple sheets, which each contain many rows, columns and cells.  You must specify which sheet you would like to manipulate.
get_sheet_names() #and
get_sheet_by_name() #help get Worksheet objects.

my_workbook.get_sheet_names()  # Result-->  ['Sheet1', 'Sheet2', 'Sheet3']

my_sheet = my_workbook.get_sheet_by_name('Sheet 1')
#verify
type(my_sheet)  # Result-->  <class 'openpyxl.worksheet.worksheet.Worksheet'>

# Cell objects have a "value" member variable with the content of that cell.
# The square brackets in sheet['A1'] get Cell objects.
# The cell() method also returns a Cell object from a sheet.
my_sheet['A1'] #will return the CELL OBJECT (NOT the value you see in the cell) created by the intersection of the first column and the first row.  The reult of sheet['A1'] will actually be:
    # Result -->  <Cell Sheet1.A1>
#so we must ASSIGN IT TO A VARIABLE before we can work with it.
my_cell = sheet['A1']
my_cell.value() # Result --> the value of the A1 cell

#Don't forget to convert the values to a string if you are working with strings
str(my_cell.value)
str(my_sheet['A1'].value)

#You can also access cell values using the .cell() method and passing arguments for the column and row:
my_sheet.cell(row=1, column=2) # Result --> <Cell Sheet1.B1>


#Python can be used to loop through the cells on your spreadsheet and manipulate data automatically.
for i in range(1, 8):
    print(i, my_sheet.cell(row = i, column =2).value)
#The code may be difficult to understand at first, but notice that in 'row = i', the 'i' will increase by 1 until it reaches 8, thus printing out each value in column 2.  The first 'i', after the 'print(' is to show the correlation between 'i' and row being printed.


# READ DATA

# WRITE TO A DATA

# UPDATE A SPREADSHEET

# SET UP A DATA STRUCTURE WITH THE UPDATED INFORMATION

# FORMULAS

# OTHER FUNCTIONALITIES

import openpyxl
my_workbook = openpyxl.Workbook()
#Review--> Determine the names of the sheets in the Excel file using the .get_sheet_names() function imported from openpyxl
my_workbook.get_sheet_names() # Result -->  ['Sheet']
#Notice the correllation between lines 144 and 146 ^^^
my_sheet = my_workbook.get_sheet_by_name('Sheet')

#How to create a NEW, EMPTY Excel document:
from openpyxl import Workbook
my_new_workbook = Workbook()
#This workbook will always be created with one sheet, which can be accessed using Workbook.active property:
my_new_worksheet = my_new_workbook.active  #  This value is set to 0 by default


wb.create_sheet() # Add a new sheet.
#<Worksheet "Sheet1">

# Once you gave a worksheet a name, you can get it as a key of the workbook:
ws3 = wb["New Title"]

# You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute
print(wb.sheetnames)  #Result-->  ['Sheet2', 'New Title', 'Sheet1']


# You can loop through worksheets
for sheet in wb:
    print(sheet.title)

# You can create copies of worksheets within a single workbook:
Workbook.copy_worksheet() method:

source = wb.active
target = wb.copy_worksheet(source)
#---> END OF REVIEW SECTION <----


#-------INSERTING ROWS & COLUMNS--------
#Much of the following info is pulled directly from the OpenPyXL official documentation

#The following openpyxl methods will allow you to insert & delete rows and columns 
openpyxl.worksheet.worksheet.Worksheet.insert_rows()
openpyxl.worksheet.worksheet.Worksheet.insert_cols()
openpyxl.worksheet.worksheet.Worksheet.delete_rows()
openpyxl.worksheet.worksheet.Worksheet.delete_cols()

#Getting a feel for the values impacted by these methods may take a couple tries :)
#For instance, the default is one row or column. To insert a row at 7 (before the existing row 7):
my_sheet.insert_rows(7)

#Deleting Rows & Columns
my_sheet.delete_cols(6, 3)

#You can also move ranges of cells within a worksheet:
my_sheet.move_range("D4:F10, rows=1, cols=2")
#This will move the cells in the range D4:F10 up one row, and right two columns. The cells will overwrite any existing cells.

#---------MANAGING SHEETS AND CELLS------------
#Review--> Access a cell value from the Excel sheet
my_sheet['A1'].value  # Result is 'None' on a black Excel document

#Review--> Change the value of a cell
my_sheet['A1'] = 37
my_sheet['A2'] = 'Pears'
#Because we know that when we create a blank Excel document, the starting value is 'None', we can clear cell values by setting them to the Python keyword: None 

#How to create a NEW, EMPTY Excel document sheet:
my_newest_worksheet1 = my_new_workbook.create_sheet("Mysheet")  # insert at the end of the sheets (default)
#or
my_newest_worksheet2 = my_new_workbook.create_sheet("Mysheet", 0)  # insert at first position
#or
my_newest_worksheet3 = my_new_workbook.create_sheet("Mysheet", -1)  # insert at penultimate position



#-------------THE OS MODULE----------------

#Review--> Files have a name and a path.
#Review--> The root folder is the lowest folder.
#Review--> In a file path, the folders and filename are separated by backslashes on #Windows and forward slashes on Linux and Mac.
#Review--> The current working directory is the folder that any relative paths are relative to.
#Use the os.path.join() function to combine folders with the correct slash.

#Begin by importing the os module to the top of your file, like so:
import os
#(openpyxl must also be imported & workbook created, but they are not in this example for the sake of simplicity).
from openpyxl.workbook import workbook

os.getcwd() #will return the current working directory.
os.chdir() #will change the current working directory.  <---- PAY SPECIAL ATTENTION TO THIS ONE

#Review--> Absolute paths begin with the root folder, relative paths do not.
#The . folder represents "this folder", the .. folder represents "the parent folder".

os.path.abspath() #returns an absolute path form of the path passed to it.
os.path.relpath() #returns the relative path between two paths passed to it.
os.makedirs() #can make folders.
os.path.getsize() #returns a file's size.
os.listdir() #returns a list of strings of filenames.
os.path.exists() #returns True if the filename passed to it exists.
os.path.isfile() #and 
os.path.isdir() #return True if they were passed a filename or file path.

#-------THE REASON WE TALKED ABOUT THE OS MODULE:  ------
#We need to use the os module to access files on our machine so that we may create, save, and delete Excel files
import os
os.chdir('/Users/tylerpritchard/Desktop/VIT/itp_week_3/day_2')
my_new_workbook.save('my_new_filename.xlsx')
#It is good to save your spreadsheets as new files, as opposed to writing over old spreadsheets and risking the loss of your work.

#OpenPyXL provides another way to save your workbook:
my_workbook.save('/Users/tylerpritchard/Desktop/VIT/itp_week_3/day_2.lecture.py')

#
#-----------------BACK TO EXCEL-------------
#
#RECAP:
#You can view and modify a sheet's name with its "title" member variable.
#Changing a cell's value is done using the square brackets, just like changing a value in a list or dictionary.
#Changes you make to the workbook object can be saved with the save() method.

# Since many office workers use Excel spreadsheets all the time, a program that can automatically edit and write Excel files could be really useful. Such a program could do the following:

# Read data from one spreadsheet and write it to parts of other spreadsheets.
# Read data from websites, text files, or the clipboard and write it to a spreadsheet.
# Automatically “clean up” data in spreadsheets. For example, it could use regular expressions to read multiple formats of phone numbers and edit them to a single, standard format.