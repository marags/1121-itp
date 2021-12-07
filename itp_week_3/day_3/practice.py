# Practice Questions:
import openpyxl
from openpyxl import Workbook

# For the following questions, imagine you have a Workbook object in the variable wb, a Worksheet object in sheet, and a Cell object in cell.
wb = Workbook()
sheet = wb['Sheet'] #Sheet is the single default sheet created with the Workbook call
c = sheet['A1']

# 1. What does the openpyxl.load_workbook() function return?
    # returns a value of the workbook data type. The workbook object represents the Excel file

# 2. What does the wb.sheetnames workbook attribute contain?
    # a list of sheet names in the workbook wb

# 3. How would you retrieve the Worksheet object for a sheet named 'Sheet1'?
wb['Sheet']

# 4. How would you retrieve the Worksheet object for the workbook’s active sheet?
wb.active 

# 5. How would you retrieve the value in the cell C5?
my_sheet = wb["Sheet"]
my_sheet['C5'].value

# 6. How would you set the value in the cell C5 to "Hello"?
my_sheet['C5'] = "Hello"

# 7. How would you retrieve the cell’s row and column as integers?

#with a sheet
print(my_sheet['C5'].row)
print(my_sheet['C5'].column)

#with a cell 
c = my_sheet['C5']

print(c.row)
print(c.column)

# 8. How would you save the workbook to the filename example.xlsx?
wb.save('path/to/exanple.xlsx')

# 9. If you needed to get the integer index for column 'M', what function would you need to call?
from openpyxl.utils import column_index_from_string
print(column_index_from_string("M"))

# 10. If you needed to get the string name for column 14, what function would you need to call?
from openpyxl.utils import get_column_letter
print(get_column_letter(14))

# BONUS: What do the sheet.max_column and sheet.max_row sheet attributes hold, and what is the data type of these attributes?
    #sheet.max_column and sheet.max_row holds integer values of the column and row for the last row and column

# BONUS: How can you retrieve a tuple of all the Cell objects from A1 to F1?
my_sheet['A1':'F1']

# BONUS: How would you set the height of row 5 to 100? 
my_sheet.row_dimensions[5].height = 100
