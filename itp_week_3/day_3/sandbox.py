import openpyxl

# filename path
filename = "/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_3/day_3/lecture.xlsx"

# load workbook
my_workbook = openpyxl.load_workbook(filename)

# get list of sheet names
#sheet_list = my_workbook.get_sheet_names() # <----- deprecated
sheet_list = my_workbook.sheetnames # updated method

print(sheet_list)

my_sheet = my_workbook.get_sheet_by_name('Sheet1')

my_sheet["A8"] = "YOU'VE"
my_sheet["B8"] = "BEEN"
my_sheet["C8"] = "HACKED"

filename_copy = "/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_3/day_3/lecture_test_save.xlsx"

my_workbook.save(filename_copy)


####----------- Day 4 work to complete day 3 -----------####

#assign workbook to an object (dictionary)
my_workbook = openpyxl.load_workbook("day_3/lecture.xlsx")

print(type(my_workbook))

print(my_workbook.get_sheet_names())

fruits = my_workbook.get_sheet_by_name("Sheet1") #function deprecated
fruits = my_workbook["Sheet1"]

print(fruits["A1"]) #prints what the cell is
print(fruits["A1"].value) #captures the value inside the cell

my_cell = fruits["B1"]
apples = my_cell.value
print(apples)

#looping through cells in a sheet
for i in range(1,8):
    print(i, fruits.cell(row = i, column = 2).value)
#alternative loop that does the same as above 
for fruit in fruits:
    print(fruit[1].value)

#create a new empty Excel document
my_workbook = openpyxl.Workbook()
my_workbook.get_sheet_names() #will return a single sheet named Sheet by default - .sheetnames is updated call
my_sheet = my_workbook.get_sheet_by_name("Sheet") #capture workseet object

#import module function 
from openpyxl import Workbook

wb = Workbook()

print("Sheets", wb.sheetnames)  #['Sheet']

my_new_worksheet = wb.active
my_sheet = wb.create_sheet() #<Worksehhet "Sheet1"> adds Sheet1 to workbook


my_sheet = wb["Sheet1"] #assigns Sheet1 to my_sheet
print(wb.sheetnames)

for sheet in wb:
     print(sheet.title)

