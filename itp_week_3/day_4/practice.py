
import openpyxl
import json
import requests

# Prepare OpenPyXl with appropriate worksheet

wb = openpyxl.load_workbook("/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_3/day_4/lecture.xlsx")

wb.create_sheet('Rick and Morty')
ws = wb['Rick and Morty']

# Headers
ws["A1"] = "Name"
ws["B1"] = "Status"
ws["C1"] = "Species"
ws["D1"] = "Gender"

# Prepare the json and requests for API calls

api_url = "https://rickandmortyapi.com/api/character?page="

row_offset = 0

for i in range(1, 43):
    api_url_iter = api_url + str(i)
    r = requests.get(api_url_iter)
    data = json.loads(r.text)

    for j in range(1, len(data['results']) + 1):
        ws["A" + str(row_offset + j+1)] = data['results'][j-1]['name']
        ws["B" + str(row_offset + j+1)] = data['results'][j-1]['status']
        ws["C" + str(row_offset + j+1)] = data['results'][j-1]['species']
        ws["D" + str(row_offset + j+1)] = data['results'][j-1]['gender']
    row_offset += 20
    
wb.save("/Users/Dev/Code/Git_Repositories/VetsInTech/Python Fundamentals/1121-itp/itp_week_3/day_4/rickandmorty.xlsx")